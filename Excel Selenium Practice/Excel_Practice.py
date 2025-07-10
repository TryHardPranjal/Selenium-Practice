import openpyxl
file_path="C://Users//pc//Downloads//download.xlsx"
book=openpyxl.load_workbook(file_path)
sheet=book.active
Dict={}
print(sheet.cell(row=3,column=4).value)

for i in range(1,sheet.max_column+1):
    if sheet.cell(row=1,column=i).value == "fruit_name":
        Dict['column']=i
    elif sheet.cell(row=1,column=i).value == "price":
        Dict['column_price']=i

for j in range(1,sheet.max_row+1):
    if sheet.cell(row=j,column=Dict['column']).value == "Apple":
        Dict["row"]=j
        break

print(Dict)
print(sheet.cell(row=Dict['row'],column=Dict['column']).value)
print(sheet.cell(row=Dict['row'],column=Dict['column_price']).value)

# for j in range(1,sheet.max_column):
#     if sheet.cell(row=1,column=j).value == "Fruit name":
#         for k in range(1,sheet.max_row):
#             if sheet.cell(row=k,column=j).value == "Apple":
#                 Dict["Row"]=k
#                 break
