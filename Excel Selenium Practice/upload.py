from email._header_value_parser import get_attribute

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

driver=webdriver.Chrome()
file="C://Users//pc//Downloads//download.xlsx"
fruit_name="Apple"
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.implicitly_wait(5)
driver.find_element(By.ID,"downloadButton").click()

def upload_excel(file_path,colName,search_term,new_Value):

    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict['column'] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == search_term:
                Dict["row"] = i
                break

    sheet.cell(row=Dict['row'],column=Dict['column']).value=new_Value
    book.save(file_path)

    # print(Dict)
    # print(sheet.cell(row=Dict['row'], column=Dict['column']).value)



#Edit excel
new_valuetoinput=945
upload_excel(file,"price","Apple",new_valuetoinput)

#upload excel
file_input=driver.find_element(By.CSS_SELECTOR,"input[type='file']")

file_input.send_keys(file)

wait=WebDriverWait(driver,5)
toast_locator=(By.CSS_SELECTOR,".Toastify__toast-body")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

#assertion
price_column=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price=int(driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text)
print(actual_price)
print(new_valuetoinput)
assert actual_price==new_valuetoinput



Answer=input("Enter A to quit: ")
if Answer=="A":
    driver.quit()